import openpyxl
import calendar
import os
import datetime
import numpy as np
from scipy.stats import norm
from scipy.linalg import cholesky
from scipy.stats import norm


def create_new_xlsx_monthly_dates(load_data, filename, secondTime = 0):

    def month_increment(start_date, num_months):
        # Calculate the new month and year
        new_month = (start_date.month + num_months - 1) % 12 + 1
        new_year = start_date.year + (start_date.month + num_months - 1) // 12
        
        # Calculate the last day of the new month
        last_day_of_month = calendar.monthrange(new_year, new_month)[1]
        
        # Ensure the new day is the last valid day of the new month if the original day doesn't exist in the new month
        new_day = min(start_date.day, last_day_of_month)
        return datetime.date(new_year, new_month, new_day)
    start_date = datetime.date(2024, 4, 30)
    monthly_dates = [month_increment(start_date, i) for i in range(load_data.shape[0])]

    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.delete_rows(1, ws.max_row)
    ws.append(['Date', '^GSPC', '^ACWX', '^GLAB.L'])

    for i, row in enumerate(load_data):
        ws.append([monthly_dates[i].strftime('%Y-%m-%d')] + row.tolist())
    wb.save(filename)

def binary_to_asset_values_qc(binary, num_assets, expected_returns, cov_matrix):
    # Convert the binary string to a list of integers
    binary_list = [int(bit) for bit in binary]
    
    # Calculate the asset values
    asset_values = []
    for i in range(num_assets):
        asset_value = expected_returns[i]
        for j in range(num_assets):
            asset_value += binary_list[j] * cov_matrix[i, j]
        asset_values.append(asset_value)
    
    return asset_values

def generate_quantum_normal_distribution_all_assets(expected_returns, cov_matrix, num_qubits, stddev):
    # Calculate the number of assets
    num_assets = len(expected_returns)
    
    # Initialize the quantum circuits
    qc_array = []
    for i in range(num_assets):
        qc = generate_quantum_normal_distribution(expected_returns[i], cov_matrix[i, i], num_qubits[i], stddev[i])
        qc_array.append(qc)
    
    return qc_array

def binary_to_asset_values_timeseries(time_series, mu, sigma):
    # Calculate the number of assets
    elements1 = [sublist[0] for sublist in time_series]
    elements2 = [sublist[1] for sublist in time_series]
    elements3 = [sublist[2] for sublist in time_series]

    # Calculate the average and standard deviation of the elements
    average1 = np.mean(elements1)
    average2 = np.mean(elements2)
    average3 = np.mean(elements3)
    stddev1 = np.std(elements1)
    stddev2 = np.std(elements2)
    stddev3 = np.std(elements3)

    # Normalize the elements
    for row in time_series:
        row[0] = ((row[0]- average1) / stddev1 ) * np.sqrt(sigma[0][0]) + mu[0]
        row[1] = ((row[1]- average2) / stddev2 ) * np.sqrt(sigma[1][1]) + mu[1]
        row[2] = ((row[2]- average3) / stddev3 ) * np.sqrt(sigma[2][2]) + mu[2]

    return time_series

def binary_to_asset_values_test(binary, num_assets, expected_returns, cov_matrix):
    # Convert the binary string to a list of integers
    binary_list = [int(bit) for bit in binary]
    
    # Calculate the asset values
    asset_values = []
    for i in range(num_assets):
        asset_value = expected_returns[i]
        for j in range(num_assets):
            asset_value += binary_list[j] * cov_matrix[i, j]
        asset_values.append(asset_value)
    
    return asset_values


def gen_data(n, mu, sigma, seed=42):
    np.random.seed(seed)
    return np.random.multivariate_normal(mu, sigma, n)

def gen_data_timeseries(n, mu, sigma, seed=42):
    np.random.seed(seed)
    return np.random.multivariate_normal(mu, sigma, n)

def gen_binary_data(n, mu, sigma, seed=42)
    data = gen_data(n, mu, sigma, seed)
    return np.array([np.array([1 if x > 0 else 0 for x in row]) for row in data])


def map_to_qubits(hamiltonian):
    # Create a SparsePauliOp using the Jordan-Wigner transform 
    mapper = JordanWignerMapper()
    mapped_hamiltonian = mapper.map(hamiltonian)

    return mapped_hamiltonian, mapper

def create_ansatze(mapped_hamiltonian, mapper):

    num_spatial_orbitals = 2
    num_particles = (1, 1)
    
    uccsd_ansatz = UCCSD(num_spatial_orbitals,
                         num_particles,
                         mapper,
                         initial_state=HartreeFock(
                             num_spatial_orbitals,
                             num_particles,
                             mapper,),)

    twolocal_ansatz = TwoLocal(mapped_hamiltonian.num_qubits,        
                           rotation_blocks=["rx", "ry"],
                           entanglement_blocks=["cz"],
                           entanglement="linear",
                           reps=2,
                           initial_state=None,)

    return uccsd_ansatz, twolocal_ansatz

def run_vqe_twolocal(backend, isa_twolocal, isa_hamiltonian_twolocal):

    num_params_twolocal = isa_twolocal.num_parameters
    x0_twolocal = np.zeros(num_params_twolocal)
    
    twolocal_dict = {
        'session_id': None,
        'iters': 0,
        'backend': backend.name,
        'maxiter': 200,
        'params': [],
        'energy': [],
        'job_ids': [],
        'job_info':[],
        'start_time': None,
        'end_time': None,
        'completed': False
    }
    
    with ibm.Session(backend=backend, max_time=3600) as session:
    
        # Create the estimator and assign the shot value
        estimator = ibm.EstimatorV2(mode=session)
        estimator.options.default_precision = 1e-1
        estimator.options.resilience_level = 1
        estimator.options.default_shots = 8192
    
        try:
            # Execute the minimization
            res = minimize(
                energy_function,
                x0_twolocal,
                args=(isa_twolocal, 
                      isa_hamiltonian_twolocal, 
                      estimator, 
                      session, 
                      twolocal_dict),
                method='COBYLA',
                tol=1e-1,
            )
    
        except ibm.IBMRuntimeError:
            print("Session closed after exceeding time limit. Stopping iteration.")
    
    return twolocal_dict


def run_vqe_uccsd(backend, isa_uccsd, isa_hamiltonian_uccsd):

    num_params_uccsd = isa_uccsd.num_parameters
    x0_uccsd = np.zeros(num_params_uccsd)
    
    uccsd_dict = {
        'session_id': None,
        'iters': 0,
        'backend': backend.name,
        'maxiter': 200,
        'params': [],
        'energy': [],
        'job_ids': [],
        'job_info':[],
        'start_time': None,
        'end_time': None,
        'completed': False
    }
    
    with ibm.Session(backend=backend, max_time=3600) as session:
    
        # Create the estimator and assign the shot value
        estimator = ibm.EstimatorV2(mode=session)
        estimator.options.default_precision = 1e-1
        estimator.options.resilience_level = 1
        estimator.options.default_shots = 8192
    
        try:
            # Execute the minimization
            res = minimize(
                energy_function,
                x0_uccsd,
                args=(isa_uccsd, 
                      isa_hamiltonian_uccsd, 
                      estimator, 
                      session, 
                      uccsd_dict),
                method='COBYLA',
                tol=1e-1,
            )
    
        except ibm.IBMRuntimeError:
            print("Session closed after exceeding time limit. Stopping iteration.")
    
    return uccsd_dict